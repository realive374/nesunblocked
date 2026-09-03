"""Write the Kahoot bulk-import spreadsheet from the shared question bank.

Kahoot's importer reads columns A-G starting at row 9, with the header on
row 8, so the layout below matches the official quiz template.
"""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from questions import VOCAB, build

HERE = os.path.dirname(os.path.abspath(__file__))
OUT = os.path.join(HERE, "Kahoot-Vocab-Quiz-Import.xlsx")

TIME_LIMIT = 20  # seconds; Kahoot allows 5, 10, 20, 30, 60, 90, 120, 240
FIRST_QUESTION_ROW = 9
HEADER_ROW = 8

HEADERS = [
    "Question - max 120 characters",
    "Answer 1 - max 75 characters",
    "Answer 2 - max 75 characters",
    "Answer 3 - max 75 characters (optional)",
    "Answer 4 - max 75 characters (optional)",
    "Time limit (sec) - 5, 10, 20, 30, 60, 90, 120 or 240",
    "Correct answer(s) - choose at least one",
]

INSTRUCTIONS = [
    VOCAB["title"] + " - Kahoot import sheet",
    "How to use: open kahoot.com, click Create > Kahoot > Import spreadsheet, then upload this file.",
    "Edit only the white cells below row 8. Keep the column order exactly as it is.",
    "Time limit must be one of: 5, 10, 20, 30, 60, 90, 120, 240.",
    'Correct answer column takes the answer number (1-4). Use "1,2" style for multiple correct answers.',
    "Example row: What does \"abstain\" mean? | obtainable | hold back | loud, energetic | to agree ... | 20 | 2",
    "",
]

BLUE = Font(name="Arial", size=10, color="0000FF")
BLACK = Font(name="Arial", size=10)
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="46178F")   # Kahoot purple
TITLE_FONT = Font(name="Arial", size=14, bold=True, color="46178F")
NOTE_FONT = Font(name="Arial", size=10, italic=True, color="595959")

wb = Workbook()
ws = wb.active
ws.title = "Sheet1"

for i, line in enumerate(INSTRUCTIONS, start=1):
    cell = ws.cell(row=i, column=1, value=line)
    cell.font = TITLE_FONT if i == 1 else NOTE_FONT
    cell.alignment = Alignment(vertical="center")

for col, header in enumerate(HEADERS, start=1):
    cell = ws.cell(row=HEADER_ROW, column=col, value=header)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
ws.row_dimensions[HEADER_ROW].height = 42

bank = build()
for offset, q in enumerate(bank):
    row = FIRST_QUESTION_ROW + offset
    values = [q["prompt"]] + q["answers"] + [TIME_LIMIT, q["correct"]]
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = BLUE if col <= 5 else BLACK
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30

note_row = FIRST_QUESTION_ROW + len(bank) + 1
note = ws.cell(row=note_row, column=1, value=(
    "Rows %d-%d each show a word and ask for its meaning. Blue text = the text Kahoot shows "
    "players. Words and definitions transcribed from the LINCS worksheets; part of speech, "
    "synonyms, sentences and pictures were left out on request."
    % (FIRST_QUESTION_ROW, FIRST_QUESTION_ROW + len(bank) - 1)))
note.font = NOTE_FONT

widths = [46, 30, 30, 30, 30, 14, 14]
for col, width in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(col)].width = width
ws.freeze_panes = "A%d" % FIRST_QUESTION_ROW

wb.save(OUT)
print("wrote", OUT, "with", len(bank), "questions")
